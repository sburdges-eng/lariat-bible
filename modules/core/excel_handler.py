"""
Excel Import/Export Handler
Handles all Excel file operations for The Lariat Bible
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any


class ExcelHandler:
    """Handles Excel import/export operations"""

    def __init__(self):
        self.workbook = None
        self.active_sheet = None

    # ========================================
    # Price Comparison Export
    # ========================================

    def export_price_comparison(self, comparison_data: List[Dict[str, Any]]) -> BytesIO:
        """
        Export price comparison data to Excel with multiple sheets

        Args:
            comparison_data: List of comparison dictionaries

        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()

        # Create main comparison sheet
        ws_main = wb.active
        ws_main.title = "Price Comparison"

        # Headers
        headers = ["Product", "Vendor A", "Price A", "Vendor B", "Price B", "Difference", "Savings %", "Recommended"]
        ws_main.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws_main[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for item in comparison_data:
            ws_main.append([
                item.get('product', ''),
                item.get('vendor_a', ''),
                item.get('price_a', 0),
                item.get('vendor_b', ''),
                item.get('price_b', 0),
                item.get('difference', 0),
                item.get('savings_percent', 0),
                item.get('recommended', ''),
            ])

        # Format columns
        ws_main.column_dimensions['A'].width = 30
        ws_main.column_dimensions['B'].width = 15
        ws_main.column_dimensions['C'].width = 12
        ws_main.column_dimensions['D'].width = 15
        ws_main.column_dimensions['E'].width = 12
        ws_main.column_dimensions['F'].width = 12
        ws_main.column_dimensions['G'].width = 12
        ws_main.column_dimensions['H'].width = 15

        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Metric", "Value"])

        total_items = len(comparison_data)
        total_savings = sum(item.get('difference', 0) for item in comparison_data)
        avg_savings_percent = sum(item.get('savings_percent', 0) for item in comparison_data) / total_items if total_items > 0 else 0

        ws_summary.append(["Total Items Compared", total_items])
        ws_summary.append(["Total Monthly Savings", f"${total_savings:.2f}"])
        ws_summary.append(["Annual Savings Potential", f"${total_savings * 12:.2f}"])
        ws_summary.append(["Average Savings Percentage", f"{avg_savings_percent:.1f}%"])
        ws_summary.append(["Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")])

        # Style summary
        for row in ws_summary.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    # ========================================
    # Order Guide Export
    # ========================================

    def export_order_guide(self, items: List[Dict[str, Any]], vendor_name: str = "Vendor") -> BytesIO:
        """
        Export order guide to Excel

        Args:
            items: List of order items
            vendor_name: Name of the vendor

        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"{vendor_name} Order Guide"

        # Headers
        headers = ["Item Code", "Description", "Pack Size", "Case Price", "Unit Price", "Unit", "Category", "Last Updated"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for item in items:
            ws.append([
                item.get('item_code', ''),
                item.get('description', ''),
                item.get('pack_size', ''),
                item.get('case_price', 0),
                item.get('unit_price', 0),
                item.get('unit', ''),
                item.get('category', ''),
                item.get('last_updated', ''),
            ])

        # Format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15

        # Format currency columns
        for row in range(2, len(items) + 2):
            ws[f'D{row}'].number_format = '$#,##0.00'
            ws[f'E{row}'].number_format = '$#,##0.00'

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    # ========================================
    # Savings Opportunities Export
    # ========================================

    def export_savings_opportunities(self, opportunities: List[Dict[str, Any]]) -> BytesIO:
        """
        Export top savings opportunities to Excel

        Args:
            opportunities: List of savings opportunities

        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Savings Opportunities"

        # Headers
        headers = ["Rank", "Product", "Current Vendor", "Current Price", "Best Vendor", "Best Price", "Savings", "Savings %", "Annual Impact"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="43E97B", end_color="43E97B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data with ranking
        for rank, item in enumerate(opportunities, 1):
            monthly_usage = item.get('monthly_usage', 10)  # Default assumption
            annual_impact = item.get('savings', 0) * monthly_usage * 12

            ws.append([
                rank,
                item.get('product', ''),
                item.get('current_vendor', ''),
                item.get('current_price', 0),
                item.get('best_vendor', ''),
                item.get('best_price', 0),
                item.get('savings', 0),
                item.get('savings_percent', 0),
                annual_impact,
            ])

            # Highlight top 3
            if rank <= 3:
                for cell in ws[rank + 1]:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        # Format columns
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15

        # Format currency and percentage columns
        for row in range(2, len(opportunities) + 2):
            ws[f'D{row}'].number_format = '$#,##0.00'
            ws[f'F{row}'].number_format = '$#,##0.00'
            ws[f'G{row}'].number_format = '$#,##0.00'
            ws[f'H{row}'].number_format = '0.0%'
            ws[f'I{row}'].number_format = '$#,##0.00'

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    # ========================================
    # Recipe Cost Export
    # ========================================

    def export_recipe_costs(self, recipes: List[Dict[str, Any]]) -> BytesIO:
        """
        Export recipe cost analysis to Excel

        Args:
            recipes: List of recipe dictionaries with cost data

        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()

        # Overview sheet
        ws_overview = wb.active
        ws_overview.title = "Recipe Overview"

        headers = ["Recipe Name", "Category", "Servings", "Ingredient Cost", "Labor Cost", "Overhead", "Total Cost", "Suggested Price", "Margin %"]
        ws_overview.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws_overview[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add recipe data
        for recipe in recipes:
            ingredient_cost = recipe.get('ingredient_cost', 0)
            labor_cost = ingredient_cost * 0.30  # 30% of ingredient cost
            overhead = ingredient_cost * 0.15  # 15% of ingredient cost
            total_cost = ingredient_cost + labor_cost + overhead
            suggested_price = total_cost / 0.55  # 45% margin
            margin_percent = 45

            ws_overview.append([
                recipe.get('name', ''),
                recipe.get('category', ''),
                recipe.get('servings', 1),
                ingredient_cost,
                labor_cost,
                overhead,
                total_cost,
                suggested_price,
                margin_percent / 100,
            ])

        # Format columns
        for col in ['D', 'E', 'F', 'G', 'H']:
            for row in range(2, len(recipes) + 2):
                ws_overview[f'{col}{row}'].number_format = '$#,##0.00'

        for row in range(2, len(recipes) + 2):
            ws_overview[f'I{row}'].number_format = '0%'

        # Create detailed ingredient sheets for each recipe
        for recipe in recipes[:5]:  # Limit to first 5 recipes for performance
            sheet_name = recipe.get('name', 'Recipe')[:31]  # Excel sheet name limit
            ws_detail = wb.create_sheet(sheet_name)

            ws_detail.append(["Ingredient", "Quantity", "Unit", "Price per Unit", "Total Cost"])

            for cell in ws_detail[1]:
                cell.fill = header_fill
                cell.font = header_font

            ingredients = recipe.get('ingredients', [])
            for ing in ingredients:
                ws_detail.append([
                    ing.get('name', ''),
                    ing.get('quantity', 0),
                    ing.get('unit', ''),
                    ing.get('price_per_unit', 0),
                    ing.get('total_cost', 0),
                ])

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    # ========================================
    # Import Functions
    # ========================================

    def import_price_list(self, file_path_or_bytes) -> pd.DataFrame:
        """
        Import price list from Excel file

        Args:
            file_path_or_bytes: File path or BytesIO object

        Returns:
            DataFrame with price data
        """
        try:
            df = pd.read_excel(file_path_or_bytes)

            # Standardize column names
            df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')

            return df
        except Exception as e:
            raise ValueError(f"Error importing price list: {str(e)}")

    def import_order_guide(self, file_path_or_bytes) -> Dict[str, Any]:
        """
        Import order guide from Excel file

        Args:
            file_path_or_bytes: File path or BytesIO object

        Returns:
            Dictionary with order guide data
        """
        try:
            df = pd.read_excel(file_path_or_bytes)

            # Convert to dictionary format
            items = df.to_dict('records')

            return {
                'items': items,
                'total_items': len(items),
                'import_date': datetime.now().isoformat(),
            }
        except Exception as e:
            raise ValueError(f"Error importing order guide: {str(e)}")

    # ========================================
    # Template Generation
    # ========================================

    def create_import_template(self, template_type: str = "price_list") -> BytesIO:
        """
        Create import template for different data types

        Args:
            template_type: Type of template (price_list, order_guide, recipe)

        Returns:
            BytesIO object containing the template
        """
        wb = Workbook()
        ws = wb.active

        if template_type == "price_list":
            ws.title = "Price List Template"
            headers = ["Item Code", "Description", "Pack Size", "Price", "Unit", "Category", "Vendor"]
            ws.append(headers)
            ws.append(["12345", "Black Pepper Fine 25lb", "25 LB", "54.99", "LB", "Spices", "Shamrock Foods"])

        elif template_type == "order_guide":
            ws.title = "Order Guide Template"
            headers = ["Item Code", "Description", "Pack Size", "Case Price", "Unit Price", "Unit", "Category"]
            ws.append(headers)
            ws.append(["12345", "Black Pepper Fine", "25 LB", "54.99", "2.20", "LB", "Spices"])

        elif template_type == "recipe":
            ws.title = "Recipe Template"
            headers = ["Recipe Name", "Ingredient", "Quantity", "Unit", "Price per Unit", "Category"]
            ws.append(headers)
            ws.append(["BBQ Sandwich", "Brisket", "8", "oz", "1.50", "Meat"])

        # Style headers
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file


# Global instance
excel_handler = ExcelHandler()
